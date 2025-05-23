import csv
import io
import json
from typing import List
import xml.etree.ElementTree as ET

import datetime

import numpy as np
from utils.date import date_to_str
from utils.exceptions import FileFormatError
from utils.timing import func_timer as timer
import openpyxl
from openpyxl.styles import Font, Alignment
import pandas
from werkzeug.datastructures import FileStorage


def format_csv(data):
    csv_string = ""
    if len(data) > 0:
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer, delimiter=",")

        writer.writerows(data)

        csv_string = csv_buffer.getvalue()

        csv_buffer.close()

    return csv_string


def dict_from_table(data, selectable_column, base_name="", nested: dict = {}):
    result = {}

    selectable_column_index = data[0].index(selectable_column)

    for row in data[1:]:
        # Dato del diccionario
        data_dict = {data[0][i]: value for i, value in enumerate(row)}

        # Nombre de la clave
        row_name: str = (
            base_name + "_" + str(row[selectable_column_index])
            if base_name
            else str(row[selectable_column_index])
        )

        result_data_dict = {}
        for d_data in data_dict:
            # Separamos el nombre de la clave por _ y nos quedamos con el primer elemento
            suffix = d_data.split("_")[0]
            value = data_dict[d_data]
            if isinstance(value, (datetime.date, datetime.datetime)):
                value = date_to_str(date=value)
            # Si el sufijo está en el diccionario nested, significa que va a pertenecer a un diccionario anidado al principal
            if suffix in nested:
                group_name = nested[suffix]
                if group_name in result_data_dict:
                    pass
                else:
                    result_data_dict[group_name] = {}

                result_data_dict[group_name][d_data.replace(f"{suffix}_", "")] = value
            else:
                result_data_dict[d_data] = value

        result[row_name] = result_data_dict

    return result


def dict_to_xml(data, root_name=None, object_name=""):
    root = ET.Element(root_name)

    def _dict_to_xml(parent, data):
        for key, value in data.items():
            if isinstance(value, dict):
                # For nested dictionaries, create a new XML element
                if isinstance(key, str) and key.startswith(object_name):
                    element_key = object_name[0 : len(object_name)]
                else:
                    element_key = str(key)
                element = ET.SubElement(parent, element_key)
                _dict_to_xml(element, value)
            else:
                # For non-dictionary values, create a new XML element and set its text
                element_key = str(key)
                ET.SubElement(parent, element_key).text = str(value)

    # Convert the dictionary to XML recursively
    _dict_to_xml(root, data)

    # Create the ElementTree from the root element
    xml_tree = ET.ElementTree(root)

    # Return the XML string representation
    return ET.tostring(root, encoding="utf-8", method="xml").decode("utf-8")


def date_serializer(obj):
    if isinstance(obj, (datetime.datetime, datetime.date)):
        return obj.isoformat()  # Convierte a string compatible con JSON
    raise TypeError(f"Tipo no serializable: {type(obj)}")


def dict_to_json(data):
    result = json.dumps(
        data,
        indent=4,
        sort_keys=False,
        default=date_serializer,
    )
    return result


def dict_to_excel(data):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Iterate through each page in the root dictionary
    for page_name, page_data in data.items():
        # Create a new worksheet for the current page
        worksheet = workbook.create_sheet(title=page_name)

        # Write column names to the first row
        column_names = list(page_data.keys())
        for col_idx, column_name in enumerate(column_names, 1):
            worksheet.cell(row=1, column=col_idx, value=column_name)

        # Iterate through the data for each page
        max_rows = max(len(col_data) for col_data in page_data.values())
        for row_idx in range(2, max_rows + 2):
            for col_idx, column_name in enumerate(column_names, 1):
                column_data = page_data[column_name]
                # Adjust for 0-based index
                value = column_data.get(row_idx - 2, "")
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

    default_sheet = workbook["Sheet"]
    workbook.remove(default_sheet)

    return workbook


def add_hyperlinks_to_excel(workbook):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Iterate through all rows in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                # Check if the cell value starts with "http"
                if cell.value and str(cell.value).startswith("http"):
                    # Create a hyperlink using the cell value
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    return workbook


def bold_column_titles_excel(workbook):
    # Iterate through all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Make the first row (header) bold
        for cell in next(sheet.iter_rows()):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    return workbook


@timer
def save_excel_to_file(workbook: openpyxl.Workbook, output_file):
    workbook.save(output_file)
    pandas_excel = pandas.read_excel(output_file, sheet_name=None)
    with pandas.ExcelWriter(output_file, engine="xlsxwriter") as writer:
        # Iterate over all sheets in the dictionary and write each sheet to the new file
        for sheet_name, df in pandas_excel.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def table_to_pandas(table: list):
    result = pandas.DataFrame(table[1:], columns=table[0]).replace({np.nan: None})
    table.clear()

    return result


def flask_csv_to_matix(file: FileStorage) -> List:
    result = []

    data = file.stream.read()
    stream = io.StringIO(data.decode("utf-8-sig"), newline=None)
    reader = csv.reader(stream)
    for row in reader:
        result.append(row)

    return result


def flask_csv_to_df(file: FileStorage) -> pandas.DataFrame:
    if len(file.name.split(".")) < 1 or file.name.split(".")[1] != "csv":
        raise FileFormatError

    matrix = flask_csv_to_matix(file)
    return table_to_pandas(matrix)


def flask_xls_to_df(file: FileStorage) -> pandas.DataFrame:
    if len(file.name.split(".")) < 1 or file.name.split(".")[1] != "xls":
        raise FileFormatError

    xls_data = pandas.read_excel(file, engine="xlrd")
    return xls_data


def enumerated_dict(iterable) -> dict:
    return {index: item for index, item in enumerate(iterable)}


def json_to_dict(json_path: str) -> dict:
    python_obj = None
    # Abrir el archivo .txt que contiene el JSON
    try:
        with open(json_path, "r") as file:
            # Cargar el contenido del archivo y convertirlo a un objeto Python
            python_obj = json.load(file)
        return python_obj
    except Exception as e:
        print("Error: ", e)


def dataframe_to_json(df, orient="index", indent=4):
    try:
        return df.to_json(orient=orient, indent=indent, force_ascii=False)
    except Exception as e:
        print("Error: ", e)


def truncate_string(s: str, length: int) -> str:
    if len(s) > length:
        return s[: length - 3] + "..."
    return s
